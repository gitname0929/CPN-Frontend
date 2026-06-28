#!/usr/bin/env python3
"""Topic benchmark task runner with streaming logs and result parsing."""

from __future__ import annotations

import json
import os
import queue
import select
import shutil
import shlex
import subprocess
import tempfile
import threading
import time
import uuid
import csv
from collections import deque
from pathlib import Path
from typing import Any

try:
    from openpyxl import load_workbook as _load_workbook
except ImportError:  # pragma: no cover
    _load_workbook = None


def _get_load_workbook():
    if _load_workbook is not None:
        return _load_workbook
    try:
        from openpyxl import load_workbook
        return load_workbook
    except ImportError as exc:
        raise RuntimeError("openpyxl is required to parse benchmark results") from exc

FEITENG_SCRIPT = os.environ.get("TOPIC1_FEITENG_SCRIPT", "/home/a/709Docker/kubernetes_py_ctl_2/run.sh")
FEITENG_WORKDIR = os.environ.get("TOPIC1_FEITENG_WORKDIR", "/home/a/709Docker/kubernetes_py_ctl_2")
ASCEND_HOST = os.environ.get("TOPIC1_ASCEND_HOST", "root@192.168.31.179")
ASCEND_SCRIPT = os.environ.get("TOPIC1_ASCEND_SCRIPT", "/root/709Docker/kubernetes_py_ctl_2/run_ascend.sh")
ASCEND_WORKDIR = os.environ.get("TOPIC1_ASCEND_WORKDIR", "/root/709Docker/kubernetes_py_ctl_2")
TOPIC1_ROUNDS = int(os.environ.get("TOPIC1_ROUNDS", "10"))
ASCEND_SSH_PASSWORD = os.environ.get("ASCEND_SSH_PASSWORD", "Mind@123")
PROJECT2_RELEASE_DIR = "/home/a/709Docker/PyTaskPredict-release"
PROJECT2_SCRIPT = f"{PROJECT2_RELEASE_DIR}/run_project2.sh"
PROJECT2_RESULTS_DIR = f"{PROJECT2_RELEASE_DIR}/results"
PROJECT3_WORKDIR = os.environ.get("PROJECT3_WORKDIR", "/home/a/709Docker/kt3_resource")
PROJECT3_ASCEND_SERVER_HOST = os.environ.get("PROJECT3_ASCEND_SERVER_HOST", "root@192.168.31.154")
PROJECT3_ASCEND_SERVER_IP = os.environ.get("PROJECT3_ASCEND_SERVER_IP", "192.168.31.154")
PROJECT3_LOCAL_FEITENG_CLIENT_IP = os.environ.get("PROJECT3_LOCAL_FEITENG_CLIENT_IP", "192.168.31.36")
PROJECT3_REMOTE_FEITENG_CLIENT_HOST = os.environ.get("PROJECT3_REMOTE_FEITENG_CLIENT_HOST", "a@192.168.31.128")
PROJECT3_REMOTE_FEITENG_CLIENT_IP = os.environ.get("PROJECT3_REMOTE_FEITENG_CLIENT_IP", "192.168.31.128")
PROJECT3_ASCEND_CLIENT1_HOST = os.environ.get("PROJECT3_ASCEND_CLIENT1_HOST", "root@192.168.31.179")
PROJECT3_ASCEND_CLIENT1_IP = os.environ.get("PROJECT3_ASCEND_CLIENT1_IP", "192.168.31.179")
PROJECT3_ASCEND_CLIENT2_HOST = os.environ.get("PROJECT3_ASCEND_CLIENT2_HOST", "root@192.168.31.247")
PROJECT3_ASCEND_CLIENT2_IP = os.environ.get("PROJECT3_ASCEND_CLIENT2_IP", "192.168.31.247")
PROJECT3_SSH_PASSWORD = os.environ.get("PROJECT3_SSH_PASSWORD", ASCEND_SSH_PASSWORD)
PROJECT3_FEITENG_SSH_PASSWORD = os.environ.get("PROJECT3_FEITENG_SSH_PASSWORD", "Sancog123")
PROJECT3_SERVER_SCRIPT = "server_cpu_extreme_compute_torch_shared.py"
PROJECT3_CLIENT_SCRIPT = "client_ascend_fix_patched.py"
PROJECT3_ASCEND_PYTHON = os.environ.get("PROJECT3_ASCEND_PYTHON", "/usr/local/miniconda3/bin/python")
PROJECT3_ASCEND_ENV = ""
PROJECT3_TAIL_BACKEND = os.environ.get("PROJECT3_TAIL_BACKEND", "cpu")
if PROJECT3_TAIL_BACKEND == "ascend":
    PROJECT3_TAIL_BACKEND = "cpu"
PROJECT3_CLIENT_BACKEND = os.environ.get("PROJECT3_CLIENT_BACKEND", "onnxruntime")
PROJECT3_FEITENG_PYTHON = os.environ.get("PROJECT3_FEITENG_PYTHON", "python3")
PROJECT3_DEFAULT_PORT = int(os.environ.get("PROJECT3_PORT", "9999"))
PROJECT3_SERVER_READY_TIMEOUT = int(os.environ.get("PROJECT3_SERVER_READY_TIMEOUT", "600"))
PROJECT3_EXCLUDE_FIRST = int(os.environ.get("PROJECT3_EXCLUDE_FIRST", "5"))
PROJECT3_DEFAULT_TASKS = int(os.environ.get("PROJECT3_DEFAULT_TASKS", "215"))
PROJECT3_RUNTIME_DIR = os.environ.get("PROJECT3_RUNTIME_DIR", f"{PROJECT3_WORKDIR}/runtime_split_cache")
PROJECT3_PSS_INTERVAL = int(os.environ.get("PROJECT3_PSS_INTERVAL", "1"))
PROJECT3_SCENARIO = "ascend_feiteng_feiteng"
PROJECT3_SCENARIO_ASCEND_ASCEND = "ascend_ascend_ascend"
PROJECT3_SCENARIOS = {PROJECT3_SCENARIO, PROJECT3_SCENARIO_ASCEND_ASCEND}
PROJECT3_BASELINE_METRICS = {
    "deit_tiny_patch16_224": {"latencyMs": 364.1976552, "memoryMb": 159},
    "levit_128": {"latencyMs": 655.9859076, "memoryMb": 175.5},
    "mobilenet_v2": {"latencyMs": 379.21, "memoryMb": 134.4},
    "mobilenet_v3_large": {"latencyMs": 294.92, "memoryMb": 144.3},
    "resnet101": {"latencyMs": 6340.610399, "memoryMb": 312.7},
    "resnet18": {"latencyMs": 932.6931026, "memoryMb": 166.8},
    "resnet50": {"latencyMs": 3203.690892, "memoryMb": 238.1},
    "mobilevgg": {"latencyMs": 2488.29769, "memoryMb": 217.8},
    "lightvgg11": {"latencyMs": 2329.589507, "memoryMb": 1154.048},
    "yolov5": {"latencyMs": 1449.505676, "memoryMb": 164.8},
    "yolov8": {"latencyMs": 572.5167201, "memoryMb": 136.3},
}
PROJECT3_ASCEND_ASCEND_BASELINE_METRICS = {
    "deit_tiny_patch16_224": {"latencyMs": 587.9247212, "memoryMb": 279.8},
    "levit_128": {"latencyMs": 267.5110607, "memoryMb": 359.7},
    "mobilenet_v2": {"latencyMs": 243.629967, "memoryMb": 248.3},
    "mobilenet_v3_large": {"latencyMs": 231.1712689, "memoryMb": 286.1},
    "resnet101": {"latencyMs": 2329.237232, "memoryMb": 1061.88},
    "resnet18": {"latencyMs": 511.0485725, "memoryMb": 166.8},
    "resnet50": {"latencyMs": 1220.934453, "memoryMb": 1110.016},
    "mobilevgg": {"latencyMs": 1462.17578, "memoryMb": 248.7},
    "lightvgg11": {"latencyMs": 1097.120133, "memoryMb": 2048},
    "yolov5": {"latencyMs": 624.2213025, "memoryMb": 358.4},
    "yolov8": {"latencyMs": 306.4463143, "memoryMb": 244.6},
}
PROJECT3_MODEL_ALIASES = {
    "deit_tiny_patch_16_224": "deit_tiny_patch16_224",
    "yolo5s": "yolov5",
    "yolov5s": "yolov5",
    "yolov5": "yolov5",
    "yolo8": "yolov8",
    "yolov8n": "yolov8",
    "yolov8": "yolov8",
    "vgg11_static": "vgg11",
}
PROJECT_ROOT = Path(__file__).resolve().parent.parent
PROJECT3_REGISTRY_PATH = Path(os.environ.get("PROJECT3_REGISTRY", str(PROJECT_ROOT / "model_registry_1v1.json")))
PROJECT3_CUT_JSON_FF_SOURCE_DIR = Path(os.environ.get("PROJECT3_CUT_JSON_FF_SOURCE_DIR", str(PROJECT_ROOT / "cut_json_ff")))
PROJECT3_CUT_JSON_SS_SOURCE_DIR = Path(os.environ.get("PROJECT3_CUT_JSON_SS_SOURCE_DIR", str(PROJECT_ROOT / "cut_json_ss")))
PROJECT4_FEITENG_WORKDIR = os.environ.get("PROJECT4_FEITENG_WORKDIR", "/home/a/7094")
PROJECT4_FEITENG_PYTHON = os.environ.get("PROJECT4_FEITENG_PYTHON", f"{PROJECT4_FEITENG_WORKDIR}/.venv/bin/python")
PROJECT4_ASCEND_HOST = os.environ.get("PROJECT4_ASCEND_HOST", "root@192.168.31.179")
PROJECT4_ASCEND_WORKDIR = os.environ.get("PROJECT4_ASCEND_WORKDIR", "/root/7094")
PROJECT4_SSH_PASSWORD = os.environ.get("PROJECT4_SSH_PASSWORD", ASCEND_SSH_PASSWORD)
PROJECT4_DEFAULT_ROUNDS = int(os.environ.get("PROJECT4_DEFAULT_ROUNDS", "300"))
PROJECT4_SCHEDULER_PORT = int(os.environ.get("PROJECT4_SCHEDULER_PORT", "9001"))
PROJECT4_WARMUP_SECONDS = int(os.environ.get("PROJECT4_WARMUP_SECONDS", "60"))
PROJECT4_USER_COUNT = int(os.environ.get("PROJECT4_USER_COUNT", "4"))
PROJECT4_DEFAULT_REQUEST_INTENSITY = int(os.environ.get("PROJECT4_DEFAULT_REQUEST_INTENSITY", "8"))
PROJECT4_LOCUST_TIMEOUT_SECONDS = int(os.environ.get("PROJECT4_LOCUST_TIMEOUT_SECONDS", "1200"))
PROJECT4_BASELINE_LRU_THROUGHPUT = float(os.environ.get("PROJECT4_BASELINE_LRU_THROUGHPUT", "0.78"))
PROJECT4_BASELINE_NO_PREDICTION_THROUGHPUT = float(os.environ.get("PROJECT4_BASELINE_NO_PREDICTION_THROUGHPUT", "0.59"))
PROJECT4_TASK_SCRIPT = r'''
import csv
import json
import os
import signal
import subprocess
import sys
import time
from pathlib import Path

base_dir = Path(__PROJECT4_BASE_DIR__)
rounds = int(__PROJECT4_ROUNDS__)
port = int(__PROJECT4_PORT__)
warmup_seconds = int(__PROJECT4_WARMUP_SECONDS__)
user_count = int(__PROJECT4_USER_COUNT__)
request_intensity = int(__PROJECT4_REQUEST_INTENSITY__)
lambda_rate = request_intensity / max(user_count, 1)
locust_timeout_seconds = int(__PROJECT4_LOCUST_TIMEOUT_SECONDS__)
baseline_lru_throughput = float(__PROJECT4_BASELINE_LRU_THROUGHPUT__)
baseline_no_prediction_throughput = float(__PROJECT4_BASELINE_NO_PREDICTION_THROUGHPUT__)
device_label = __PROJECT4_DEVICE_LABEL__
scheduler_url = f"http://127.0.0.1:{port}"
k8s_dir = base_dir / "k8s"
trace_file = base_dir / "fixed_trace_300.csv"
locust_file = base_dir / "LocustFixedTrace.py"
scheduler_python = base_dir / ".venv" / "bin" / "python"
locust_bin = base_dir / ".venv" / "bin" / "locust"
scheduler_file = base_dir / "orchestrator" / "our_scheduler_api.py"
locust_csv_prefix = base_dir / "cache_test_result"
model_deployments = [
    "resnet18-deploy",
    "resnet50-deploy",
    "resnet101-deploy",
    "mobilevgg-deploy",
    "lightvgg11-deploy",
    "levit128-deploy",
    "deit-tiny-patch16-224-deploy",
    "mobilenetv2-deploy",
    "mobilenetv3-deploy",
    "yolov8n-deploy",
    "yolov5s-deploy",
    "hetero-task-deploy",
]


def is_scheduler_ready():
    try:
        result = subprocess.run(
            ["curl", "-s", f"{scheduler_url}/health"],
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
            text=True,
            timeout=5,
            cwd=str(base_dir),
        )
        return result.returncode == 0 and '"ok":true' in result.stdout.replace(" ", "")
    except Exception:
        return False


def run_cmd(cmd, timeout=60):
    try:
        result = subprocess.run(
            cmd,
            cwd=str(base_dir),
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
            text=True,
            timeout=timeout,
        )
    except subprocess.TimeoutExpired as exc:
        return {
            "cmd": " ".join(str(part) for part in cmd),
            "returncode": 124,
            "stdout": (exc.stdout or "")[-3000:] if isinstance(exc.stdout, str) else "",
            "stderr": f"command timed out after {timeout}s",
        }
    return {
        "cmd": " ".join(str(part) for part in cmd),
        "returncode": result.returncode,
        "stdout": (result.stdout or "")[-3000:],
        "stderr": (result.stderr or "")[-3000:],
    }


def stop_existing_scheduler():
    stopped = []
    scheduler_path = str(scheduler_file)
    current_pid = os.getpid()
    for item in Path("/proc").iterdir():
        if not item.name.isdigit():
            continue
        try:
            pid = int(item.name)
            if pid == current_pid:
                continue
            raw = (item / "cmdline").read_bytes()
        except Exception:
            continue
        cmdline = raw.replace(b"\x00", b" ").decode("utf-8", errors="ignore")
        if scheduler_path in cmdline:
            try:
                os.kill(pid, signal.SIGTERM)
                stopped.append(pid)
            except ProcessLookupError:
                pass
    if stopped:
        time.sleep(2)
    print(f"已清理残留调度器进程: {len(stopped)} 个", flush=True)


def clean_old_locust_csv():
    for suffix in ["_stats.csv", "_stats_history.csv", "_failures.csv", "_exceptions.csv"]:
        path = Path(str(locust_csv_prefix) + suffix)
        if path.exists():
            path.unlink()
    print("已清理旧 Locust 结果文件", flush=True)


def redeploy_k8s_resources():
    if not k8s_dir.exists():
        raise RuntimeError(f"k8s 目录不存在: {k8s_dir}")
    result = run_cmd(["k3s", "kubectl", "apply", "-f", str(k8s_dir)], timeout=180)
    if result["returncode"] != 0:
        raise RuntimeError(result["stderr"] or result["stdout"] or "k8s apply failed")
    print("Kubernetes 资源已重新部署", flush=True)


def check_model_deployments_exist():
    for dep in model_deployments:
        result = run_cmd(["k3s", "kubectl", "scale", "deploy", dep, "--replicas=0", "-n", "default"], timeout=60)
        # keep cluster state predictable for later checks


def clean_pods():
    failures = []
    for dep in model_deployments:
        result = run_cmd(["k3s", "kubectl", "scale", "deploy", dep, "--replicas=0", "-n", "default"], timeout=60)
        if result["returncode"] != 0:
            failures.append(f"{dep}: {result['stderr'] or result['stdout']}")
    if failures:
        raise RuntimeError("模型 Deployment 缩容失败: " + "; ".join(failures[:3]))
    print("已将模型与异构任务 Deployment 缩容到 0", flush=True)


def assert_deployments_exist():
    missing = []
    for dep in model_deployments:
        result = run_cmd(["k3s", "kubectl", "get", "deploy", dep, "-n", "default"], timeout=30)
        if result["returncode"] != 0:
            missing.append(dep)
    if missing:
        raise RuntimeError("模型 Deployment 不存在: " + ", ".join(missing))
    print("模型 Deployment 检查通过", flush=True)


def start_cache_algorithm():
    if is_scheduler_ready():
        print(f"复用已运行的缓存调度服务: {scheduler_url}", flush=True)
        return None
    if not scheduler_file.exists():
        raise RuntimeError(f"scheduler file not found: {scheduler_file}")
    python_bin = str(scheduler_python if scheduler_python.exists() else Path(sys.executable))
    log_file = base_dir / "project4_scheduler.log"
    out = log_file.open("a", encoding="utf-8")
    process = subprocess.Popen(
        [python_bin, str(scheduler_file)],
        cwd=str(base_dir),
        stdout=out,
        stderr=subprocess.STDOUT,
        text=True,
        env={**os.environ, "PYTHONUNBUFFERED": "1"},
        start_new_session=True,
    )
    for _ in range(60):
        if is_scheduler_ready():
            print(f"缓存调度服务已启动: {scheduler_url}", flush=True)
            return process
        if process.poll() is not None:
            raise RuntimeError(f"scheduler exited with code {process.returncode}")
        time.sleep(1)
    raise RuntimeError("scheduler health check timeout")


def reset_scheduler_state():
    result = run_cmd(["curl", "-s", "-X", "POST", f"{scheduler_url}/reset"], timeout=30)
    print("已请求重置缓存调度器状态", flush=True)
    return result


def check_scheduler_health():
    result = run_cmd(["curl", "-s", f"{scheduler_url}/health"], timeout=10)
    if result["returncode"] != 0:
        raise RuntimeError(result["stderr"] or result["stdout"] or "scheduler health failed")
    print("缓存调度服务健康检查通过", flush=True)


def get_scheduler_total_requests():
    result = run_cmd(["curl", "-s", "--max-time", "3", f"{scheduler_url}/metrics"], timeout=5)
    if result["returncode"] != 0:
        return None
    try:
        data = json.loads(result["stdout"])
        return int((data.get("stats") or {}).get("total_requests", 0))
    except Exception:
        return None


def analyze_trace():
    if not trace_file.exists():
        raise RuntimeError(f"请求流文件不存在: {trace_file}")
    rows = []
    with trace_file.open("r", encoding="utf-8") as file_obj:
        reader = csv.DictReader(file_obj)
        for row in reader:
            rows.append(row)

    model_tasks = []
    hetero_tasks = []
    task_distribution = {}

    for row in rows:
        task_group = row.get("task_group", "")
        model_id = row.get("model_id", "")
        task_type = row.get("task_type", "")

        if task_group == "intelligent_model":
            task_name = model_id
            if task_name and task_name not in model_tasks:
                model_tasks.append(task_name)
        elif task_group == "heterogeneous_perception_interaction":
            task_name = task_type
            if task_name and task_name not in hetero_tasks:
                hetero_tasks.append(task_name)
        else:
            task_name = model_id or task_type

        if task_name:
            task_distribution[task_name] = task_distribution.get(task_name, 0) + 1

    return {
        "traceFile": str(trace_file),
        "totalRequests": len(rows),
        "supportedTaskCount": len(model_tasks) + len(hetero_tasks),
        "modelTaskCount": len(model_tasks),
        "heteroTaskCount": len(hetero_tasks),
        "modelTasks": model_tasks,
        "heteroTasks": hetero_tasks,
        "taskDistribution": task_distribution,
    }


def print_trace_analysis(analysis):
    print(f"请求流文件: {analysis['traceFile']}", flush=True)
    print(f"请求流总数: {analysis['totalRequests']} 条；支持任务数: {analysis['supportedTaskCount']} 个", flush=True)
    print(f"智能模型任务: {analysis['modelTaskCount']} 个 - {', '.join(analysis['modelTasks']) or '无'}", flush=True)
    print(f"异构感知交互任务: {analysis['heteroTaskCount']} 个 - {', '.join(analysis['heteroTasks']) or '无'}", flush=True)
    if analysis["taskDistribution"]:
        distribution = ", ".join(f"{name}:{count}" for name, count in analysis["taskDistribution"].items())
        print(f"请求流分布: {distribution}", flush=True)


def run_locust_test():
    if not trace_file.exists():
        raise RuntimeError(f"请求流文件不存在: {trace_file}")
    if not locust_file.exists():
        raise RuntimeError(f"Locust 文件不存在: {locust_file}")
    locust_exe = str(locust_bin if locust_bin.exists() else "locust")
    cmd = [
        locust_exe,
        "-f", str(locust_file),
        "--host", scheduler_url,
        "--headless",
        "-u", str(user_count),
        "-r", "1",
        "--csv", str(locust_csv_prefix),
        "--only-summary",
    ]
    env = {**os.environ, "LAMBDA_RATE": str(lambda_rate)}
    stdout_file = base_dir / "cache_test_locust_stdout.log"
    stderr_file = base_dir / "cache_test_locust_stderr.log"
    start_time = time.time()
    reached_target = False
    with stdout_file.open("w", encoding="utf-8") as out, stderr_file.open("w", encoding="utf-8") as err:
        process = subprocess.Popen(
            cmd,
            cwd=str(base_dir),
            env=env,
            stdout=out,
            stderr=err,
            text=True,
            start_new_session=True,
        )
        while True:
            elapsed = time.time() - start_time
            total_requests = get_scheduler_total_requests()
            if total_requests is not None:
                shown = min(total_requests, rounds)
                print(f"已执行 {shown}/{rounds}", flush=True)
                if total_requests >= rounds:
                    reached_target = True
                    break
            if process.poll() is not None:
                break
            if elapsed > locust_timeout_seconds:
                break
            time.sleep(2)
        if process.poll() is None:
            try:
                os.killpg(os.getpgid(process.pid), signal.SIGINT)
                process.wait(timeout=20)
            except Exception:
                try:
                    os.killpg(os.getpgid(process.pid), signal.SIGKILL)
                except Exception:
                    pass
    stdout_text = stdout_file.read_text(encoding="utf-8", errors="ignore")[-3000:] if stdout_file.exists() else ""
    stderr_text = stderr_file.read_text(encoding="utf-8", errors="ignore")[-3000:] if stderr_file.exists() else ""
    if not reached_target:
        raise RuntimeError(stderr_text or stdout_text or "Locust 未在规定时间内完成目标请求数")
    print(f"Locust 压测完成，目标请求数 {rounds}", flush=True)


def parse_locust_result():
    stats_file = Path(str(locust_csv_prefix) + "_stats.csv")
    history_file = Path(str(locust_csv_prefix) + "_stats_history.csv")
    if not stats_file.exists():
        raise RuntimeError(f"Locust 结果文件不存在: {stats_file}")
    aggregated = None
    with stats_file.open("r", encoding="utf-8") as file_obj:
        for row in csv.DictReader(file_obj):
            if row.get("Name") == "Aggregated":
                aggregated = row
                break
    if aggregated is None:
        raise RuntimeError("未找到 Locust Aggregated 统计行")
    avg_throughput = float(aggregated.get("Requests/s", 0) or 0)
    steady_rps_values = []
    if history_file.exists():
        with history_file.open("r", encoding="utf-8") as file_obj:
            rows = list(csv.DictReader(file_obj))
        start_ts = None
        for row in rows:
            if row.get("Name") != "Aggregated":
                continue
            try:
                ts = float(row.get("Timestamp", 0) or 0)
                rps = float(row.get("Requests/s", 0) or 0)
            except Exception:
                continue
            if start_ts is None:
                start_ts = ts
            if ts - start_ts >= warmup_seconds:
                steady_rps_values.append(rps)
    steady_throughput = round(sum(steady_rps_values) / len(steady_rps_values), 2) if steady_rps_values else None
    throughput = steady_throughput if steady_throughput is not None else round(avg_throughput, 2)
    return {
        "throughput": throughput,
        "avgThroughput": round(avg_throughput, 2),
        "steadyThroughput": steady_throughput,
        "warmupSeconds": warmup_seconds,
    }


def build_algorithm_result(ours_throughput):
    ours = round(float(ours_throughput), 2)
    lru = round(float(baseline_lru_throughput), 2)
    no_prediction = round(float(baseline_no_prediction_throughput), 2)
    return {
        "lru": lru,
        "noPrediction": no_prediction,
        "ours": ours,
        "improvementVsLru": round((ours - lru) / lru * 100, 2) if lru > 0 else None,
        "improvementVsNoPrediction": round((ours - no_prediction) / no_prediction * 100, 2) if no_prediction > 0 else None,
    }


started_process = None
started_by_task = False
try:
    request_rate = request_intensity
    clean_old_locust_csv()
    stop_existing_scheduler()
    redeploy_k8s_resources()
    time.sleep(5)
    assert_deployments_exist()
    clean_pods()
    time.sleep(3)
    assert_deployments_exist()
    started_process = start_cache_algorithm()
    if started_process is not None:
        started_by_task = True
    time.sleep(2)
    reset_scheduler_state()
    check_scheduler_health()
    trace_analysis = analyze_trace()
    print_trace_analysis(trace_analysis)
    print(f"计划执行 {rounds} 轮；前 {warmup_seconds}s 不计入吞吐率", flush=True)
    print(f"正在以 {request_rate:.0f} req/s 输入强度回放固定 {rounds} 个请求流", flush=True)
    run_locust_test()
    parsed = parse_locust_result()
    algorithm_result = build_algorithm_result(parsed["throughput"])
    result = {
        "device": device_label,
        "rounds": rounds,
        "requestIntensity": request_intensity,
        "traceAnalysis": trace_analysis,
        "algorithmResult": algorithm_result,
        "executed": rounds,
        "warmupSeconds": parsed["warmupSeconds"],
        "throughput": parsed["throughput"],
        "avgThroughput": parsed["avgThroughput"],
        "steadyThroughput": parsed["steadyThroughput"],
    }
    print("__PROJECT4_RESULT__" + json.dumps(result, ensure_ascii=False), flush=True)
finally:
    if started_by_task and started_process and started_process.poll() is None:
        try:
            os.killpg(os.getpgid(started_process.pid), signal.SIGTERM)
            started_process.wait(timeout=10)
        except Exception:
            try:
                os.killpg(os.getpgid(started_process.pid), signal.SIGKILL)
            except Exception:
                pass
'''
PROJECT3_REGISTRY_MODEL_ALIASES = {
    "yolov5": "yolo5",
    "yolov8": "yolo8",
}
PROJECT3_SERVER_ARCH_ALIASES = {
    "lightvgg11": "lightvgg",
    "yolo8": "yolov8",
    "yolov8n": "yolov8",
}
SSH_BASE_OPTS = [
    "-o",
    "StrictHostKeyChecking=no",
    "-o",
    "UserKnownHostsFile=/dev/null",
    "-o",
    "ConnectTimeout=15",
    "-o",
    "PreferredAuthentications=password",
    "-o",
    "PubkeyAuthentication=no",
]

TASK_RETENTION_SEC = 3600
TOPIC1_ASCEND_IDLE_TIMEOUT_SEC = int(os.environ.get("TOPIC1_ASCEND_IDLE_TIMEOUT_SEC", "900"))
TOPIC1_ASCEND_TOTAL_TIMEOUT_SEC = int(os.environ.get("TOPIC1_ASCEND_TOTAL_TIMEOUT_SEC", "7200"))
_tasks: dict[str, dict[str, Any]] = {}
_lock = threading.RLock()


def _now_ms() -> int:
    return int(time.time() * 1000)


def _cleanup_old_tasks() -> None:
    cutoff = time.time() - TASK_RETENTION_SEC
    with _lock:
        stale = [task_id for task_id, task in _tasks.items() if task["created_at"] < cutoff]
        for task_id in stale:
            _tasks.pop(task_id, None)


def _format_number(value: Any) -> Any:
    if value is None:
        return None
    try:
        number = float(value)
    except (TypeError, ValueError):
        return value
    rounded = round(number, 2)
    return int(rounded) if rounded == int(rounded) else rounded


def _normalize_models(models: list[str]) -> str:
    cleaned = [item.strip() for item in models if item and item.strip()]
    if not cleaned or cleaned == ["all"]:
        return ""
    return ",".join(cleaned)


def _has_sshpass() -> bool:
    return subprocess.run(
        ["which", "sshpass"],
        capture_output=True,
        text=True,
        check=False,
    ).returncode == 0


def _build_ssh_command(remote_cmd: str) -> list[str]:
    ssh_cmd = ["ssh", *SSH_BASE_OPTS, ASCEND_HOST, remote_cmd]
    if not _has_sshpass():
        return ssh_cmd
    return ["sshpass", "-p", ASCEND_SSH_PASSWORD, *ssh_cmd]


def _build_scp_command(remote_path: str, local_path: str) -> list[str]:
    scp_cmd = ["scp", *SSH_BASE_OPTS, f"{ASCEND_HOST}:{remote_path}", local_path]
    if not _has_sshpass():
        return scp_cmd
    return ["sshpass", "-p", ASCEND_SSH_PASSWORD, *scp_cmd]


def _build_scp_upload_command_for(local_path: str, host: str, remote_path: str, password: str | None = None) -> list[str]:
    scp_cmd = ["scp", *SSH_BASE_OPTS, local_path, f"{host}:{remote_path}"]
    if not _has_sshpass():
        return scp_cmd
    return ["sshpass", "-p", password or ASCEND_SSH_PASSWORD, *scp_cmd]


def _build_ssh_command_for(host: str, remote_cmd: str, password: str | None = None) -> list[str]:
    ssh_cmd = ["ssh", *SSH_BASE_OPTS, host, f"bash -lc {shlex.quote(remote_cmd)}"]
    if not _has_sshpass():
        return ssh_cmd
    return ["sshpass", "-p", password or ASCEND_SSH_PASSWORD, *ssh_cmd]


def _build_ssh_raw_command_for(host: str, remote_cmd: str, password: str | None = None) -> list[str]:
    ssh_cmd = ["ssh", *SSH_BASE_OPTS, host, remote_cmd]
    if not _has_sshpass():
        return ssh_cmd
    return ["sshpass", "-p", password or ASCEND_SSH_PASSWORD, *ssh_cmd]


def _build_run_command(platform: str, rounds: int, models: list[str]) -> tuple[list[str], Path, str | None]:
    models_arg = _normalize_models(models)
    output_name = f"all_models_state_benchmark_round{rounds}.xlsx"

    if platform == "feiteng":
        cmd = ["bash", FEITENG_SCRIPT, "--rounds", str(rounds)]
        if models_arg:
            cmd.extend(["--models", models_arg])
        return cmd, Path(FEITENG_WORKDIR) / output_name, FEITENG_WORKDIR

    if platform == "ascend":
        remote_cmd = (
            f"cd {shlex.quote(ASCEND_WORKDIR)} && "
            f"bash {shlex.quote(ASCEND_SCRIPT)} --rounds {int(rounds)}"
        )
        if models_arg:
            remote_cmd += f" --models {shlex.quote(models_arg)}"
        cmd = _build_ssh_command(remote_cmd)
        remote_output = f"{ASCEND_WORKDIR}/{output_name}"
        return cmd, Path(remote_output), None

    raise ValueError(f"Unsupported platform: {platform}")


def _validate_topic1_ascend_runtime() -> None:
    check_cmd = (
        f"test -f {shlex.quote(ASCEND_SCRIPT)} && "
        f"! grep -q 'phytium-infer' {shlex.quote(ASCEND_SCRIPT)}"
    )
    completed = subprocess.run(_build_ssh_command(check_cmd), capture_output=True, text=True, timeout=30)
    if completed.returncode == 0:
        return
    output = (completed.stderr or completed.stdout or "").strip()
    raise RuntimeError(
        "课题一昇腾启动脚本校验失败：当前远端脚本不存在，或仍引用飞腾镜像 phytium-infer。"
        f"请在 {ASCEND_HOST}:{ASCEND_SCRIPT} 配置昇腾专用脚本，或通过 TOPIC1_ASCEND_SCRIPT/TOPIC1_ASCEND_WORKDIR 指向正确目录。"
        + (f" 远端输出: {output}" if output else "")
    )


def _build_project2_run_command(board: str, dataset_group: str, load_level: str) -> tuple[list[str], Path, str | None]:
    output_name = f"{board}_{dataset_group}_{load_level}.csv"
    cmd = [
        "bash",
        PROJECT2_SCRIPT,
        "--board",
        board,
        "--dataset_group",
        dataset_group,
        "--load_level",
        load_level,
    ]
    return cmd, Path(PROJECT2_RESULTS_DIR) / output_name, PROJECT2_RELEASE_DIR


def _project4_device_label(device: str) -> str:
    return "昇腾" if device == "ascend" else "飞腾"


def _project4_script(base_dir: str, rounds: int, device_label: str, request_intensity: int) -> str:
    return (
        PROJECT4_TASK_SCRIPT
        .replace("__PROJECT4_BASE_DIR__", repr(base_dir))
        .replace("__PROJECT4_ROUNDS__", repr(int(rounds)))
        .replace("__PROJECT4_PORT__", repr(PROJECT4_SCHEDULER_PORT))
        .replace("__PROJECT4_WARMUP_SECONDS__", repr(PROJECT4_WARMUP_SECONDS))
        .replace("__PROJECT4_USER_COUNT__", repr(PROJECT4_USER_COUNT))
        .replace("__PROJECT4_REQUEST_INTENSITY__", repr(int(request_intensity)))
        .replace("__PROJECT4_LOCUST_TIMEOUT_SECONDS__", repr(PROJECT4_LOCUST_TIMEOUT_SECONDS))
        .replace("__PROJECT4_BASELINE_LRU_THROUGHPUT__", repr(PROJECT4_BASELINE_LRU_THROUGHPUT))
        .replace("__PROJECT4_BASELINE_NO_PREDICTION_THROUGHPUT__", repr(PROJECT4_BASELINE_NO_PREDICTION_THROUGHPUT))
        .replace("__PROJECT4_DEVICE_LABEL__", repr(device_label))
    )


def _build_project4_run_command(device: str, rounds: int) -> tuple[list[str], Path, str | None]:
    device_label = _project4_device_label(device)
    request_intensity = PROJECT4_DEFAULT_REQUEST_INTENSITY
    if device == "feiteng":
        script = _project4_script(PROJECT4_FEITENG_WORKDIR, rounds, device_label, request_intensity)
        python_bin = PROJECT4_FEITENG_PYTHON if Path(PROJECT4_FEITENG_PYTHON).exists() else "python3"
        return [python_bin, "-u", "-c", script], Path(PROJECT4_FEITENG_WORKDIR) / "cache_algorithm_result.json", PROJECT4_FEITENG_WORKDIR
    if device == "ascend":
        script = _project4_script(PROJECT4_ASCEND_WORKDIR, rounds, device_label, request_intensity)
        remote_cmd = f"cd {shlex.quote(PROJECT4_ASCEND_WORKDIR)} && ./.venv/bin/python -u -c {shlex.quote(script)}"
        return _build_ssh_raw_command_for(PROJECT4_ASCEND_HOST, remote_cmd, PROJECT4_SSH_PASSWORD), Path(PROJECT4_ASCEND_WORKDIR) / "cache_algorithm_result.json", None
    raise ValueError(f"Unsupported project4 device: {device}")


def _parse_project4_result(lines: list[str]) -> list[dict[str, Any]]:
    marker = "__PROJECT4_RESULT__"
    for line in reversed(lines):
        if marker not in line:
            continue
        raw = line.split(marker, 1)[1].strip()
        data = json.loads(raw)
        return [
            {
                "执行设备": data.get("device"),
                "执行总轮数": data.get("rounds"),
                "请求强度(req/s)": data.get("requestIntensity"),
                "请求流总数": (data.get("traceAnalysis") or {}).get("totalRequests"),
                "智能模型任务数": (data.get("traceAnalysis") or {}).get("modelTaskCount"),
                "异构感知任务数": (data.get("traceAnalysis") or {}).get("heteroTaskCount"),
                "吞吐量(req/s)": data.get("throughput"),
                "LRU吞吐量(req/s)": (data.get("algorithmResult") or {}).get("lru"),
                "Ours w/o prediction吞吐量(req/s)": (data.get("algorithmResult") or {}).get("noPrediction"),
                "相较LRU提升比(%)": (data.get("algorithmResult") or {}).get("improvementVsLru"),
                "相较Ours w/o prediction提升比(%)": (data.get("algorithmResult") or {}).get("improvementVsNoPrediction"),
            }
        ]
    raise RuntimeError("未找到课题四执行结果标记")


def _append_log(task: dict[str, Any], message: str, log_type: str = "stdout") -> None:
    entry = {"time": _now_ms(), "message": message.rstrip("\n"), "type": log_type}
    task["logs"].append(entry)
    for subscriber in list(task["subscribers"]):
        try:
            subscriber.put_nowait(entry)
        except queue.Full:
            pass


def _terminate_process(process: subprocess.Popen, wait_seconds: int = 10) -> None:
    if process.poll() is not None:
        return
    process.terminate()
    try:
        process.wait(timeout=wait_seconds)
    except subprocess.TimeoutExpired:
        process.kill()
        process.wait()


def _topic1_timeout_message(payload: dict[str, Any], idle_seconds: int, total_seconds: int) -> str | None:
    if payload.get("topicId", 1) != 1 or payload.get("platform") != "ascend":
        return None
    if idle_seconds >= TOPIC1_ASCEND_IDLE_TIMEOUT_SEC:
        return (
            f"课题一昇腾执行已 {idle_seconds}s 无新日志，已自动终止。"
            "当前通常卡在远端 Docker 镜像加载/导入阶段，请在 192.168.31.179 检查 Docker 状态、磁盘空间和镜像包完整性。"
        )
    if total_seconds >= TOPIC1_ASCEND_TOTAL_TIMEOUT_SEC:
        return f"课题一昇腾执行超过总超时 {TOPIC1_ASCEND_TOTAL_TIMEOUT_SEC}s，已自动终止。"
    return None


def _topic1_failure_message(payload: dict[str, Any], output_lines: list[str], return_code: int) -> str:
    if payload.get("topicId", 1) == 1 and payload.get("platform") == "ascend":
        output_text = "\n".join(output_lines[-80:])
        if "ErrImagePull" in output_text or "ImagePullBackOff" in output_text:
            if "registry-1.docker.io" in output_text or "Docker Hub" in output_text or "Client.Timeout" in output_text:
                return (
                    "课题一昇腾 Pod 镜像拉取失败：远端节点访问 Docker Hub 超时。"
                    "请在 192.168.31.179/集群节点预拉取相关镜像，或配置可访问的镜像仓库/加速器后重试。"
                )
            return "课题一昇腾 Pod 镜像拉取失败，请检查远端镜像名称、仓库权限和节点镜像缓存。"
    return f"脚本退出码 {return_code}"


def _parse_summary_xlsx(xlsx_path: Path) -> list[dict[str, Any]]:
    load_workbook = _get_load_workbook()

    workbook = load_workbook(xlsx_path, read_only=True, data_only=True)
    sheet = workbook["summary"]
    rows = list(sheet.iter_rows(values_only=True))
    workbook.close()
    if len(rows) < 2:
        return []

    headers = [str(item).strip() if item is not None else "" for item in rows[0]]
    header_index = {name: idx for idx, name in enumerate(headers)}
    required = ["model", "avg_cold_start_sec", "avg_warm_start_sec", "theta_percent"]
    for name in required:
        if name not in header_index:
            raise RuntimeError(f"Missing column in summary sheet: {name}")

    parsed: list[dict[str, Any]] = []
    for row in rows[1:]:
        if not row or row[header_index["model"]] in (None, ""):
            continue
        parsed.append(
            {
                "model": row[header_index["model"]],
                "avgColdToHot": _format_number(row[header_index["avg_cold_start_sec"]]),
                "avgWarmToHot": _format_number(row[header_index["avg_warm_start_sec"]]),
                "improvementRatio": _format_number(row[header_index["theta_percent"]]),
            }
        )
    return parsed


def _parse_csv_rows(csv_path: Path) -> list[dict[str, Any]]:
    with csv_path.open("r", encoding="utf-8-sig", newline="") as file_obj:
        reader = csv.DictReader(file_obj)
        rows: list[dict[str, Any]] = []
        for row in reader:
            parsed = {}
            for key, value in row.items():
                if key is None:
                    continue
                parsed[key] = _format_number(value) if value not in (None, "") else value
            if any(value not in (None, "") for value in parsed.values()):
                rows.append(parsed)
        return rows


def _project3_model_name(model: str) -> str:
    cleaned = (model or "resnet50").strip()
    return PROJECT3_MODEL_ALIASES.get(cleaned, cleaned)


def _project3_pth_path(model: str) -> str:
    return f"{PROJECT3_WORKDIR}/models/backbone_{model}_pretrained.pth"


def _project3_cut_json_path(model: str, scenario: str) -> str:
    cut_model = PROJECT3_REGISTRY_MODEL_ALIASES.get(model, model)
    if scenario == PROJECT3_SCENARIO_ASCEND_ASCEND:
        dirname = "cut_json_ss"
        prefix = "cut_ss"
    else:
        dirname = "cut_json_ff"
        prefix = "cut_ff"
    path = Path(PROJECT3_WORKDIR) / dirname / f"{prefix}_{cut_model}.json"
    if model == "lightvgg11" and not path.exists():
        path = Path(PROJECT3_WORKDIR) / dirname / f"{prefix}_lightvg11.json"
    return str(path)


def _project3_scenario_label(scenario: str) -> str:
    labels = {
        PROJECT3_SCENARIO: "昇腾 server / 飞腾 client x2",
        PROJECT3_SCENARIO_ASCEND_ASCEND: "昇腾 server / 昇腾 client x2",
    }
    return labels.get(scenario, scenario)


def _project3_scenario_cut_dir(scenario: str) -> str:
    return "cut_json_ss" if scenario == PROJECT3_SCENARIO_ASCEND_ASCEND else "cut_json_ff"


def _project3_cut_source_dir(dirname: str) -> Path:
    if dirname == "cut_json_ss":
        return PROJECT3_CUT_JSON_SS_SOURCE_DIR
    return PROJECT3_CUT_JSON_FF_SOURCE_DIR


def _load_project3_registry() -> dict[str, Any]:
    if not PROJECT3_REGISTRY_PATH.exists():
        return {}
    with PROJECT3_REGISTRY_PATH.open("r", encoding="utf-8") as file_obj:
        data = json.load(file_obj)
    return data if isinstance(data, dict) else {}


def _project3_registry_key(model: str) -> str:
    return PROJECT3_REGISTRY_MODEL_ALIASES.get(model, model)


def _project3_server_arch(arch: str) -> str:
    normalized = (arch or "").strip()
    return PROJECT3_SERVER_ARCH_ALIASES.get(normalized, normalized)


def _project3_model_config(model: str, scenario: str) -> dict[str, Any]:
    registry = _load_project3_registry()
    registry_key = _project3_registry_key(model)
    item = registry.get(registry_key) if registry else None
    if isinstance(item, dict):
        if item.get("enabled") is False:
            reason = item.get("disabled_reason") or "disabled in registry"
            raise RuntimeError(f"model {registry_key} is disabled: {reason}")
        pth = item.get("pth") or _project3_pth_path(model)
        arch = _project3_server_arch(str(item.get("arch") or model))
    else:
        pth = _project3_pth_path(model)
        arch = _project3_server_arch(model)
    cut_json = _project3_cut_json_path(model, scenario)
    device_keys = ["device_0", "device_1"]
    if not Path(str(pth)).exists():
        raise RuntimeError(f"pth file not found for {registry_key}: {pth}")
    if not Path(str(cut_json)).exists():
        profile = "ss" if scenario == PROJECT3_SCENARIO_ASCEND_ASCEND else "ff"
        raise RuntimeError(f"cut_json not found for {registry_key}/{profile}: {cut_json}")
    if scenario == PROJECT3_SCENARIO_ASCEND_ASCEND:
        clients = [
            {"type": "ascend-remote", "device_key": "device_0", "ip": PROJECT3_ASCEND_CLIENT1_IP, "host": PROJECT3_ASCEND_CLIENT1_HOST},
            {"type": "ascend-remote", "device_key": "device_1", "ip": PROJECT3_ASCEND_CLIENT2_IP, "host": PROJECT3_ASCEND_CLIENT2_HOST},
        ]
    else:
        clients = [
            {"type": "ft-local", "device_key": "device_0", "ip": PROJECT3_LOCAL_FEITENG_CLIENT_IP},
            {"type": "ft-remote", "device_key": "device_1", "ip": PROJECT3_REMOTE_FEITENG_CLIENT_IP, "host": PROJECT3_REMOTE_FEITENG_CLIENT_HOST},
        ]
    return {
        "registryKey": registry_key,
        "pth": str(pth),
        "cutJson": str(cut_json),
        "arch": str(arch),
        "numClasses": None,
        "deviceKey": "device_0",
        "deviceKeys": device_keys,
        "clients": clients,
        "runtimeDir": PROJECT3_RUNTIME_DIR,
    }


def _project3_baseline_path(model: str, scenario: str) -> Path:
    del scenario
    return Path(PROJECT3_WORKDIR) / "baseline_ff_reserved" / f"baseline_{model}.json"


def _json_get_path(data: Any, path: list[Any]) -> Any:
    cur = data
    for key in path:
        if isinstance(key, int):
            if not isinstance(cur, list) or len(cur) <= key:
                return None
            cur = cur[key]
            continue
        if not isinstance(cur, dict) or key not in cur:
            return None
        cur = cur[key]
    return cur


def _parse_project3_baseline(model: str, scenario: str) -> dict[str, Any]:
    metrics_map = PROJECT3_ASCEND_ASCEND_BASELINE_METRICS if scenario == PROJECT3_SCENARIO_ASCEND_ASCEND else PROJECT3_BASELINE_METRICS
    metrics = metrics_map.get(model) or {}
    source = "内置双昇腾全本地 baseline 表" if scenario == PROJECT3_SCENARIO_ASCEND_ASCEND else "内置双飞腾全本地 baseline 表"
    return {
        "latencyMs": _format_number(metrics.get("latencyMs")),
        "memoryMb": _format_number(metrics.get("memoryMb")),
        "source": source,
    }


def _parse_project3_memory_metrics(model: str, scenario: str) -> dict[str, Any]:
    metrics_map = PROJECT3_ASCEND_ASCEND_BASELINE_METRICS if scenario == PROJECT3_SCENARIO_ASCEND_ASCEND else PROJECT3_BASELINE_METRICS
    baseline = metrics_map.get(model) or {}
    source = "内置双昇腾全本地 baseline 表" if scenario == PROJECT3_SCENARIO_ASCEND_ASCEND else "内置双飞腾全本地 baseline 表"
    return {"baselineMemoryMb": _format_number(baseline.get("memoryMb")), "methodMemoryMb": None, "source": source}


def _format_project3_ratio(value: Any) -> Any:
    try:
        num = float(value)
    except (TypeError, ValueError):
        return None
    if abs(num) <= 1:
        num *= 100
    return _format_number(num)


def _extract_metric(text: str, name: str) -> Any:
    import re

    pattern = re.compile(rf"{re.escape(name)}=([-+]?\d+(?:\.\d+)?|nan|null|None)")
    matches = pattern.findall(text or "")
    for raw in reversed(matches):
        if raw in ("nan", "null", "None"):
            continue
        try:
            return _format_number(float(raw))
        except ValueError:
            continue
    return None


def _terminate_process(process: subprocess.Popen | None) -> None:
    if not process or process.poll() is not None:
        return
    process.terminate()
    try:
        process.wait(timeout=10)
    except subprocess.TimeoutExpired:
        process.kill()


def _wait_for_project3_server_port(
    task: dict[str, Any],
    port: int,
    server_process: subprocess.Popen | None = None,
    server_lines: list[str] | None = None,
    timeout_sec: int | None = None,
) -> bool:
    timeout_sec = timeout_sec or PROJECT3_SERVER_READY_TIMEOUT
    deadline = time.time() + timeout_sec
    remote_cmd = f"ss -lnt | grep -q ':{int(port)} '"
    cmd = _build_ssh_command_for(PROJECT3_ASCEND_SERVER_HOST, remote_cmd, PROJECT3_SSH_PASSWORD)
    next_log_at = time.time()
    while time.time() < deadline:
        if server_process is not None and server_process.poll() is not None:
            return False
        completed = subprocess.run(cmd, capture_output=True, text=True, timeout=15)
        has_current_server_ready_log = server_lines is None or any(
            "Listening" in line and f":{int(port)}" in line for line in server_lines
        )
        if completed.returncode == 0 and has_current_server_ready_log:
            return True
        if time.time() >= next_log_at:
            remaining = max(0, int(deadline - time.time()))
            ready_reason = "端口未监听" if completed.returncode != 0 else "等待本次 server Listening 日志"
            _append_log(
                task,
                f"server 仍在启动，继续等待端口 {port}，剩余 {remaining}s；{ready_reason}；首次运行可能正在导出 ONNX 缓存",
                "info",
            )
            next_log_at = time.time() + 10
        time.sleep(2)
    return False


def _cleanup_project3_server_port(port: int) -> None:
    port_value = int(port)
    remote_cmd = f'''
port={port_value}
collect_pids() {{
  {{ ss -lntp "sport = :$port" 2>/dev/null | sed -n 's/.*pid=\([0-9][0-9]*\).*/\1/p'; }} || true
  {{ command -v lsof >/dev/null 2>&1 && lsof -tiTCP:"$port" -sTCP:LISTEN 2>/dev/null; }} || true
  {{ command -v fuser >/dev/null 2>&1 && fuser -n tcp "$port" 2>/dev/null; }} || true
}}
pkill -TERM -f 'Ascend/.*/bin/atc' || true
pkill -TERM -f 'Ascend/.*/bin/atc.bin' || true
pkill -TERM -f ' ccec ' || true
pkill -TERM -f {shlex.quote(f'{PROJECT3_SERVER_SCRIPT}.*--port {port_value}')} || true
for attempt in 1 2 3 4 5; do
  pids=$(collect_pids | tr ' ' '\n' | sed '/^$/d' | sort -u)
  if [ -z "$pids" ]; then
    exit 0
  fi
  for pid in $pids; do kill -TERM "$pid" 2>/dev/null || true; done
  sleep 1
done
pids=$(collect_pids | tr ' ' '\n' | sed '/^$/d' | sort -u)
for pid in $pids; do kill -KILL "$pid" 2>/dev/null || true; done
sleep 1
pids=$(collect_pids | tr ' ' '\n' | sed '/^$/d' | sort -u)
if [ -n "$pids" ]; then
  echo "port $port still occupied by: $pids" >&2
  exit 1
fi
'''.strip()
    subprocess.run(
        _build_ssh_command_for(PROJECT3_ASCEND_SERVER_HOST, remote_cmd, PROJECT3_SSH_PASSWORD),
        capture_output=True,
        text=True,
        timeout=30,
    )


def _project3_server_port_in_use(port: int) -> bool:
    remote_cmd = f"ss -lnt 'sport = :{int(port)}' 2>/dev/null | grep -q ':{int(port)} '"
    completed = subprocess.run(
        _build_ssh_command_for(PROJECT3_ASCEND_SERVER_HOST, remote_cmd, PROJECT3_SSH_PASSWORD),
        capture_output=True,
        text=True,
        timeout=15,
    )
    return completed.returncode == 0


def _start_logged_process(
    task: dict[str, Any],
    label: str,
    cmd: list[str],
    workdir: str | None = None,
) -> tuple[subprocess.Popen, list[str]]:
    _append_log(task, f"{label} 命令: {' '.join(shlex.quote(part) for part in cmd)}", "info")
    process = subprocess.Popen(
        cmd,
        cwd=workdir,
        stdout=subprocess.PIPE,
        stderr=subprocess.STDOUT,
        text=True,
        bufsize=1,
        env={**os.environ, "PYTHONUNBUFFERED": "1"},
    )
    lines: list[str] = []

    def _reader() -> None:
        assert process.stdout is not None
        for line in iter(process.stdout.readline, ""):
            clean = line.rstrip("\n")
            lines.append(clean)
            _append_log(task, f"[{label}] {clean}", "stdout")

    threading.Thread(target=_reader, daemon=True).start()
    return process, lines


def _project3_server_command(model: str, scenario: str, tasks: int, port: int) -> list[str]:
    config = _project3_model_config(model, scenario)
    device_key = config["deviceKey"]
    device_keys = config.get("deviceKeys") or [device_key]
    clients = config.get("clients") or []
    total_tasks = int(tasks) + PROJECT3_EXCLUDE_FIRST
    # ATC is disabled for project 3. Do not source Ascend toolkit here because it
    # exposes atc/ccec and can reactivate ONNX->OM paths in legacy scripts.
    env_prefix = ""
    num_classes_arg = ""
    if config.get("numClasses") not in (None, "", 0):
        num_classes_arg = f"--num-classes {int(config['numClasses'])} "
    command = (
        f"cd {shlex.quote(PROJECT3_WORKDIR)} && "
        f"{env_prefix}"
        f"{shlex.quote(PROJECT3_ASCEND_PYTHON)} -u {shlex.quote(PROJECT3_SERVER_SCRIPT)} "
        f"--host 0.0.0.0 --port {int(port)} "
        f"--pth {shlex.quote(config['pth'])} "
        f"--cut-json {shlex.quote(config['cutJson'])} "
        f"--arch {shlex.quote(config['arch'])} "
        f"{num_classes_arg}"
        f"--device-key {shlex.quote(device_key)} --device-keys {shlex.quote(','.join(device_keys))} "
        f"--tail-backend {shlex.quote(PROJECT3_TAIL_BACKEND)} --device-id 0 "
        f"--expect-clients {max(1, len(clients))} --tasks-per-client {total_tasks} "
        f"--exclude-first {PROJECT3_EXCLUDE_FIRST} --require-hello "
        f"--runtime-dir {shlex.quote(config['runtimeDir'])} "
        f"--log-every 10 --print-avg-every 50"
    )
    return _build_ssh_command_for(PROJECT3_ASCEND_SERVER_HOST, command, PROJECT3_SSH_PASSWORD)


def _project3_client_command_for(
    client_type: str,
    device_key: str,
    tasks: int,
    port: int,
    model: str,
    host: str | None = None,
    password: str | None = None,
) -> tuple[list[str], str | None]:
    python_bin = PROJECT3_ASCEND_PYTHON if client_type.startswith("ascend") else PROJECT3_FEITENG_PYTHON
    command = [
        python_bin,
        "-u",
        PROJECT3_CLIENT_SCRIPT,
        "--server-ip",
        PROJECT3_ASCEND_SERVER_IP,
        "--port",
        str(port),
        "--workdir",
        "./model_cache",
        "--device-key",
        device_key,
        "--num-tasks",
        str(tasks),
        "--exclude-first",
        str(PROJECT3_EXCLUDE_FIRST),
    ]
    command.append("--force-cpu")
    safe_model = "".join(ch if ch.isalnum() or ch in "_.-" else "_" for ch in model)
    safe_client_type = "".join(ch if ch.isalnum() or ch in "_.-" else "_" for ch in client_type)
    pss_csv = f"{PROJECT3_WORKDIR}/client_pss_{safe_model}_{safe_client_type}_{device_key}.csv"
    client_cmd = " ".join(shlex.quote(part) for part in command)
    wrapper = f'''
cd {shlex.quote(PROJECT3_WORKDIR)}
rm -f {shlex.quote(pss_csv)}
printf 'timestamp,datetime,pid,pss_kb,rss_kb,pss_mb,rss_mb\n' > {shlex.quote(pss_csv)}
({client_cmd}) &
CLIENT_PID=$!
(
  while kill -0 "$CLIENT_PID" 2>/dev/null; do
    TS=$(date +%s)
    TIME_STR=$(date "+%Y-%m-%dT%H:%M:%S%z")
    if [ -r "/proc/$CLIENT_PID/smaps_rollup" ]; then
      PSS_KB=$(awk '/^Pss:/ {{print $2; exit}}' "/proc/$CLIENT_PID/smaps_rollup")
      RSS_KB=$(awk '/^Rss:/ {{print $2; exit}}' "/proc/$CLIENT_PID/smaps_rollup")
    elif [ -r "/proc/$CLIENT_PID/smaps" ]; then
      PSS_KB=$(awk '/^Pss:/ {{sum += $2}} END {{print sum + 0}}' "/proc/$CLIENT_PID/smaps")
      RSS_KB=$(awk '/^Rss:/ {{sum += $2}} END {{print sum + 0}}' "/proc/$CLIENT_PID/smaps")
    else
      sleep {shlex.quote(str(PROJECT3_PSS_INTERVAL))}
      continue
    fi
    PSS_MB=$(awk -v kb="$PSS_KB" 'BEGIN {{printf "%.3f", kb / 1024}}')
    RSS_MB=$(awk -v kb="$RSS_KB" 'BEGIN {{printf "%.3f", kb / 1024}}')
    echo "$TS,$TIME_STR,$CLIENT_PID,$PSS_KB,$RSS_KB,$PSS_MB,$RSS_MB" >> {shlex.quote(pss_csv)}
    sleep {shlex.quote(str(PROJECT3_PSS_INTERVAL))}
  done
) &
MONITOR_PID=$!
wait "$CLIENT_PID"
CLIENT_CODE=$?
kill "$MONITOR_PID" 2>/dev/null || true
PSS_AVG=$(awk -F, 'NR>1 {{sum += $6; count++}} END {{if (count) printf "%.3f", sum / count; else printf ""}}' {shlex.quote(pss_csv)})
PSS_PEAK=$(awk -F, 'NR>1 {{if ($6 > max) max = $6; count++}} END {{if (count) printf "%.3f", max; else printf ""}}' {shlex.quote(pss_csv)})
PSS_COUNT=$(awk -F, 'NR>1 {{count++}} END {{print count + 0}}' {shlex.quote(pss_csv)})
echo CLIENT_PSS_AVG_MB=$PSS_AVG
echo CLIENT_PSS_PEAK_MB=$PSS_PEAK
echo CLIENT_PSS_SAMPLES=$PSS_COUNT
exit "$CLIENT_CODE"
'''.strip()
    if client_type in ("ft-remote", "ascend-remote"):
        remote_host = host or (PROJECT3_REMOTE_FEITENG_CLIENT_HOST if client_type == "ft-remote" else PROJECT3_ASCEND_CLIENT1_HOST)
        remote_password = password or (PROJECT3_FEITENG_SSH_PASSWORD if client_type == "ft-remote" else PROJECT3_SSH_PASSWORD)
        return _build_ssh_command_for(remote_host, wrapper, remote_password), None
    return ["bash", "-lc", wrapper], None


def _project3_client_commands(scenario: str, model: str, tasks: int, port: int) -> list[tuple[str, list[str], str | None]]:
    config = _project3_model_config(model, scenario)
    clients = config.get("clients") or []
    commands: list[tuple[str, list[str], str | None]] = []
    for index, client in enumerate(clients, start=1):
        client_type = str(client.get("type") or "ft-local")
        device_key = str(client.get("device_key") or config["deviceKey"])
        ip = str(client.get("ip") or "")
        host = str(client.get("host") or "") or None
        password = PROJECT3_FEITENG_SSH_PASSWORD if client_type == "ft-remote" else PROJECT3_SSH_PASSWORD
        cmd, workdir = _project3_client_command_for(client_type, device_key, tasks, port, model, host, password)
        commands.append((f"client{index}:{client_type}:{device_key}:{ip}", cmd, workdir))
    if commands:
        return commands
    return []


def _sync_project3_file_to_host(task: dict[str, Any], host: str, filename: str, password: str | None = None) -> None:
    local_path = Path(PROJECT3_WORKDIR) / filename
    remote_path = f"{PROJECT3_WORKDIR}/{filename}"
    if not local_path.exists():
        raise RuntimeError(f"课题三本地脚本不存在，无法同步: {local_path}")
    mkdir_cmd = _build_ssh_command_for(host, f"mkdir -p {shlex.quote(PROJECT3_WORKDIR)}", password or PROJECT3_SSH_PASSWORD)
    completed = subprocess.run(mkdir_cmd, capture_output=True, text=True, timeout=30)
    if completed.returncode != 0:
        raise RuntimeError(f"创建远程目录失败 {host}: {completed.stderr.strip() or completed.stdout.strip()}")
    cmd = _build_scp_upload_command_for(str(local_path), host, remote_path, password or PROJECT3_SSH_PASSWORD)
    completed = subprocess.run(cmd, capture_output=True, text=True, timeout=60)
    if completed.returncode != 0:
        raise RuntimeError(
            f"同步课题三脚本到 {host} 失败: {completed.stderr.strip() or completed.stdout.strip()}"
        )
    _append_log(task, f"已同步课题三脚本到 {host}: {filename}", "info")


def _sync_project3_dir_to_host(task: dict[str, Any], host: str, dirname: str, password: str | None = None) -> None:
    local_path = _project3_cut_source_dir(dirname) if dirname in ("cut_json_ff", "cut_json_ss") else Path(PROJECT3_WORKDIR) / dirname
    if not local_path.exists() or not local_path.is_dir():
        raise RuntimeError(f"课题三本地目录不存在，无法同步: {local_path}")
    mkdir_cmd = _build_ssh_command_for(host, f"mkdir -p {shlex.quote(PROJECT3_WORKDIR)}", password or PROJECT3_SSH_PASSWORD)
    completed = subprocess.run(mkdir_cmd, capture_output=True, text=True, timeout=30)
    if completed.returncode != 0:
        raise RuntimeError(f"创建远程目录失败 {host}: {completed.stderr.strip() or completed.stdout.strip()}")
    remote_path = f"{host}:{PROJECT3_WORKDIR}/"
    scp_cmd = ["scp", "-r", *SSH_BASE_OPTS, str(local_path), remote_path]
    if _has_sshpass():
        scp_cmd = ["sshpass", "-p", password or PROJECT3_SSH_PASSWORD, *scp_cmd]
    completed = subprocess.run(scp_cmd, capture_output=True, text=True, timeout=300)
    if completed.returncode != 0:
        raise RuntimeError(f"同步课题三目录到 {host} 失败: {completed.stderr.strip() or completed.stdout.strip()}")
    _append_log(task, f"已同步课题三目录到 {host}: {dirname}", "info")


def _sync_project3_local_cut_json(task: dict[str, Any], dirname: str) -> None:
    source_dir = _project3_cut_source_dir(dirname)
    if not source_dir.exists() or not source_dir.is_dir():
        raise RuntimeError(f"{dirname} 源目录不存在: {source_dir}")
    target = Path(PROJECT3_WORKDIR) / dirname
    target.parent.mkdir(parents=True, exist_ok=True)
    if target.exists():
        shutil.rmtree(target)
    shutil.copytree(source_dir, target)
    _append_log(task, f"已同步 {dirname} 到本机运行目录: {target}", "info")


def _sync_project3_runtime_scripts(task: dict[str, Any], scenario: str) -> None:
    cut_dir = _project3_scenario_cut_dir(scenario)
    _sync_project3_local_cut_json(task, cut_dir)
    _sync_project3_file_to_host(task, PROJECT3_ASCEND_SERVER_HOST, PROJECT3_SERVER_SCRIPT, PROJECT3_SSH_PASSWORD)
    _sync_project3_dir_to_host(task, PROJECT3_ASCEND_SERVER_HOST, cut_dir, PROJECT3_SSH_PASSWORD)
    if scenario == PROJECT3_SCENARIO_ASCEND_ASCEND:
        for host in (PROJECT3_ASCEND_CLIENT1_HOST, PROJECT3_ASCEND_CLIENT2_HOST):
            _sync_project3_file_to_host(task, host, PROJECT3_CLIENT_SCRIPT, PROJECT3_SSH_PASSWORD)
            _sync_project3_dir_to_host(task, host, cut_dir, PROJECT3_SSH_PASSWORD)
    else:
        _sync_project3_file_to_host(task, PROJECT3_REMOTE_FEITENG_CLIENT_HOST, PROJECT3_CLIENT_SCRIPT, PROJECT3_FEITENG_SSH_PASSWORD)
        _sync_project3_dir_to_host(task, PROJECT3_REMOTE_FEITENG_CLIENT_HOST, cut_dir, PROJECT3_FEITENG_SSH_PASSWORD)


def _project3_improvement(baseline_value: Any, ours_value: Any) -> Any:
    try:
        baseline_num = float(baseline_value)
        ours_num = float(ours_value)
    except (TypeError, ValueError):
        return None
    if baseline_num <= 0:
        return None
    return _format_number((baseline_num - ours_num) / baseline_num * 100)


def _project3_result_row(
    scenario: str,
    scenario_label: str,
    model: str,
    baseline: dict[str, Any],
    memory_metrics: dict[str, Any],
    latency=None,
    client_latencies: dict[str, Any] | None = None,
    pss_metrics: dict[str, Any] | None = None,
) -> dict[str, Any]:
    pss_metrics = pss_metrics or {}
    baseline_latency = baseline.get("latencyMs")
    baseline_memory = memory_metrics.get("baselineMemoryMb") or baseline.get("memoryMb")
    pss_peak_avg = pss_metrics.get("pssPeakAvgMb")
    return {
        "baseline平均延迟(ms)": baseline_latency,
        "模型分割后平均E2E(ms)": latency,
        "任务完成时间性能提升(%)": _project3_improvement(baseline_latency, latency),
        "baseline占用内存(MB)": baseline_memory,
        "模型分割后占用内存(MB)": pss_peak_avg,
        "内存资源节省比(%)": _project3_improvement(baseline_memory, pss_peak_avg),
        "任务完成时间性能提升指标": "10%",
        "内存资源节省比指标": "20%",
    }


def _project3_row_has_required_metrics(row: dict[str, Any]) -> bool:
    required_keys = [key for key in row if key in ("模型分割后平均E2E(ms)", "模型分割后占用内存(MB)")]
    return all(row.get(key) not in (None, "") for key in required_keys)


def _run_project3_model(
    task: dict[str, Any],
    scenario: str,
    model: str,
    tasks: int,
    port: int,
) -> dict[str, Any]:
    scenario_label = _project3_scenario_label(scenario)
    baseline = _parse_project3_baseline(model, scenario)
    memory_metrics = _parse_project3_memory_metrics(model, scenario)
    server_process: subprocess.Popen | None = None
    client_processes: list[subprocess.Popen] = []
    client_outputs: list[tuple[str, list[str]]] = []
    server_lines: list[str] = []
    pss_metrics: dict[str, Any] | None = None

    try:
        _append_log(
            task,
            f"开始执行模型 ({scenario_label} / {model} / 统计 {tasks} 次 / 预热 {PROJECT3_EXCLUDE_FIRST} 次)",
            "info",
        )
        _append_log(task, f"Baseline 数据: {baseline['source']}", "info")
        _append_log(task, f"内存数据: {memory_metrics['source']}", "info")
        model_config = _project3_model_config(model, scenario)
        _append_log(
            task,
            "课题三配置: "
            f"registry={model_config['registryKey']} arch={model_config['arch']} "
            f"device_keys={','.join(model_config.get('deviceKeys') or [model_config['deviceKey']])} "
            f"clients={model_config.get('clients')} tail_backend={PROJECT3_TAIL_BACKEND} "
            f"client_backend={PROJECT3_CLIENT_BACKEND} "
            f"cut_json={model_config['cutJson']} runtime_dir={model_config['runtimeDir']}",
            "info",
        )
        _cleanup_project3_server_port(port)
        if _project3_server_port_in_use(port):
            _append_log(task, f"server 端口 {port} 仍被占用，重试清理", "info")
            _cleanup_project3_server_port(port)
        if _project3_server_port_in_use(port):
            raise RuntimeError(f"server 端口仍被占用: {PROJECT3_ASCEND_SERVER_IP}:{port}")
        server_cmd = _project3_server_command(model, scenario, tasks, port)
        server_process, server_lines = _start_logged_process(task, f"server:{model}", server_cmd)
        task["process"] = server_process

        _append_log(task, f"等待 server 端口就绪: {PROJECT3_ASCEND_SERVER_IP}:{port}", "info")
        if not _wait_for_project3_server_port(task, port, server_process, server_lines):
            if server_process.poll() is not None:
                raise RuntimeError(f"server 启动失败，退出码 {server_process.returncode}")
            raise RuntimeError(f"server 端口未就绪: {PROJECT3_ASCEND_SERVER_IP}:{port}")
        _append_log(task, "server 端口已就绪", "success")

        for label, client_cmd, client_workdir in _project3_client_commands(scenario, model, tasks, port):
            client_process, client_lines = _start_logged_process(task, f"{label}:{model}", client_cmd, client_workdir)
            client_processes.append(client_process)
            client_outputs.append((label, client_lines))
        if client_processes:
            task["process"] = client_processes[-1]

        failed_clients = []
        for index, process in enumerate(client_processes):
            client_code = process.wait()
            if client_code != 0:
                failed_clients.append(f"{client_outputs[index][0]}={client_code}")
        if failed_clients:
            _append_log(task, f"client 脚本退出码: {', '.join(failed_clients)}，缺失指标将保留为空", "error")

        try:
            server_process.wait(timeout=30)
        except subprocess.TimeoutExpired:
            _terminate_process(server_process)
            _cleanup_project3_server_port(port)
        server_code = server_process.poll()
        if server_code not in (0, None):
            _append_log(task, f"server 脚本退出码 {server_code}", "error")

        latencies = []
        client_latencies: dict[str, Any] = {}
        pss_peak_values = []
        for _label, lines in client_outputs:
            client_text = "\n".join(lines)
            parts = _label.split(":")
            client_type = parts[1] if len(parts) >= 2 else _label
            client_metric_key = client_type
            if client_type == "ascend-remote" and parts and parts[0].startswith("client"):
                client_metric_key = f"{client_type}-{parts[0].replace('client', '')}"
            latency_item = _extract_metric(client_text, "avg_e2e_ms")
            if latency_item is None:
                latency_item = _extract_metric(client_text, "e2e_ms")
            if latency_item is not None:
                try:
                    latencies.append(float(latency_item))
                    client_latencies[client_metric_key] = _format_number(float(latency_item))
                except (TypeError, ValueError):
                    pass
            pss_item = _extract_metric(client_text, "CLIENT_PSS_PEAK_MB")
            if pss_item is not None:
                try:
                    pss_num = float(pss_item)
                    pss_peak_values.append(pss_num)
                except (TypeError, ValueError):
                    pass
        latency = _format_number(sum(latencies) / len(latencies)) if latencies else None
        pss_metrics = {
            "pssPeakAvgMb": _format_number(sum(pss_peak_values) / len(pss_peak_values)) if pss_peak_values else None,
        }
        return _project3_result_row(scenario, scenario_label, model, baseline, memory_metrics, latency, client_latencies, pss_metrics)
    except Exception as exc:
        for process in client_processes:
            _terminate_process(process)
        _terminate_process(server_process)
        _cleanup_project3_server_port(port)
        row = _project3_result_row(scenario, scenario_label, model, baseline, memory_metrics, pss_metrics=pss_metrics)
        if not _project3_row_has_required_metrics(row):
            row["错误"] = str(exc)
        _append_log(task, f"模型 {model} 执行失败: {exc}", "error")
        return row


def _run_project3_task(task_id: str) -> None:
    with _lock:
        task = _tasks.get(task_id)
    if not task:
        return

    payload = task["payload"]
    scenario = payload.get("scenario") or PROJECT3_SCENARIO
    if scenario not in PROJECT3_SCENARIOS:
        scenario = PROJECT3_SCENARIO
    raw_models = payload.get("models") or [payload.get("model") or "resnet50"]
    models = [_project3_model_name(item) for item in raw_models if item]
    if not models:
        models = ["resnet50"]
    tasks = int(payload.get("numTasks") or PROJECT3_DEFAULT_TASKS)
    port = int(payload.get("port") or PROJECT3_DEFAULT_PORT)
    scenario_label = _project3_scenario_label(scenario)
    _append_log(task, f"开始执行课题三 ({scenario_label} / {len(models)} 个模型 / {tasks} 次)", "info")
    _sync_project3_runtime_scripts(task, scenario)
    rows = []
    for index, model in enumerate(models, start=1):
        _append_log(task, f"[{index}/{len(models)}] 准备执行模型: {model}", "info")
        rows.append(_run_project3_model(task, scenario, model, tasks, port))

    has_error = any(row.get("错误") and not _project3_row_has_required_metrics(row) for row in rows)
    with _lock:
        current = _tasks.get(task_id)
        if not current:
            return
        current["result_rows"] = rows
        current["status"] = "failed" if has_error else "completed"
        current["error"] = "部分模型执行失败" if has_error else None
        current["process"] = None
        _append_log(current, f"执行完成，共 {len(rows)} 条结果", "success" if not has_error else "error")


def _fetch_remote_xlsx(remote_path: str) -> Path:
    temp_file = Path(tempfile.mkstemp(prefix="topic1_result_", suffix=".xlsx")[1])
    cmd = _build_scp_command(remote_path, str(temp_file))
    completed = subprocess.run(cmd, capture_output=True, text=True, timeout=120)
    if completed.returncode != 0:
        temp_file.unlink(missing_ok=True)
        raise RuntimeError(completed.stderr.strip() or "Failed to download remote result file")
    return temp_file


def _resolve_result_rows(task: dict[str, Any]) -> list[dict[str, Any]]:
    payload = task["payload"]
    if payload.get("topicId") == 2:
        _, output_path, _ = _build_project2_run_command(
            payload["board"],
            payload["dataset_group"],
            payload["load_level"],
        )
        if not output_path.exists():
            raise RuntimeError(f"Result file not found: {output_path}")
        return _parse_csv_rows(output_path)

    _, output_path, workdir = _build_run_command(
        payload["platform"],
        TOPIC1_ROUNDS,
        payload["models"],
    )

    local_path = output_path
    temp_path: Path | None = None
    try:
        if payload["platform"] == "ascend":
            temp_path = _fetch_remote_xlsx(str(output_path))
            local_path = temp_path
        elif not local_path.exists():
            raise RuntimeError(f"Result file not found: {local_path}")

        return _parse_summary_xlsx(local_path)
    finally:
        if temp_path is not None:
            temp_path.unlink(missing_ok=True)


def _run_process(task_id: str) -> None:
    with _lock:
        task = _tasks.get(task_id)
    if not task:
        return

    payload = task["payload"]
    topic_id = payload.get("topicId", 1)
    output_lines: list[str] = []
    try:
        if topic_id == 2:
            cmd, output_path, workdir = _build_project2_run_command(
                payload["board"],
                payload["dataset_group"],
                payload["load_level"],
            )
        elif topic_id == 4:
            cmd, output_path, workdir = _build_project4_run_command(
                payload["device"],
                payload["rounds"],
            )
        else:
            cmd, output_path, workdir = _build_run_command(
                payload["platform"],
                TOPIC1_ROUNDS,
                payload["models"],
            )
            if payload["platform"] == "ascend":
                _validate_topic1_ascend_runtime()
    except Exception as exc:
        with _lock:
            task = _tasks.get(task_id)
            if task:
                task["status"] = "failed"
                task["error"] = str(exc)
                _append_log(task, str(exc), "error")
        return

    if topic_id == 2:
        board_label = "昇腾" if payload["board"] == "st" else "飞腾"
        _append_log(
            task,
            f"开始执行课题二 ({board_label} / {payload['dataset_group']} / {payload['load_level']})",
            "info",
        )
    elif topic_id == 4:
        device_label = _project4_device_label(payload["device"])
        target_label = "192.168.31.179" if payload["device"] == "ascend" else "本机"
        _append_log(task, f"开始执行课题四 ({device_label} / {target_label} / 总轮数 {payload['rounds']} / 请求强度 {PROJECT4_DEFAULT_REQUEST_INTENSITY} req/s)", "info")
    else:
        platform_label = "飞腾" if payload["platform"] == "feiteng" else "昇腾"
        _append_log(task, f"开始执行课题一 ({platform_label})", "info")
    _append_log(task, f"命令: {' '.join(shlex.quote(part) for part in cmd)}", "info")
    _append_log(task, f"结果文件: {output_path}", "info")

    if topic_id == 2:
        Path(PROJECT2_RESULTS_DIR).mkdir(parents=True, exist_ok=True)

    try:
        process = subprocess.Popen(
            cmd,
            cwd=workdir,
            stdout=subprocess.PIPE,
            stderr=subprocess.STDOUT,
            text=True,
            bufsize=1,
            env={**os.environ, "PYTHONUNBUFFERED": "1"},
        )
    except Exception as exc:
        with _lock:
            task = _tasks.get(task_id)
            if task:
                task["status"] = "failed"
                task["error"] = str(exc)
                _append_log(task, f"启动失败: {exc}", "error")
        return

    with _lock:
        task = _tasks.get(task_id)
        if task:
            task["process"] = process

    assert process.stdout is not None
    start_time = time.time()
    last_output_time = start_time
    timed_out_message: str | None = None
    while process.poll() is None:
        ready, _, _ = select.select([process.stdout], [], [], 1)
        if not ready:
            now = time.time()
            timed_out_message = _topic1_timeout_message(payload, int(now - last_output_time), int(now - start_time))
            if timed_out_message:
                _terminate_process(process)
                break
            continue

        line = process.stdout.readline()
        if not line:
            break
        last_output_time = time.time()
        with _lock:
            current = _tasks.get(task_id)
            if not current:
                _terminate_process(process)
                return
            output_lines.append(line.rstrip("\n"))
            _append_log(current, line.rstrip("\n"), "stdout")

    for line in process.stdout.readlines():
        output_lines.append(line.rstrip("\n"))
        with _lock:
            current = _tasks.get(task_id)
            if current:
                _append_log(current, line.rstrip("\n"), "stdout")

    return_code = process.wait()
    with _lock:
        task = _tasks.get(task_id)
        if not task:
            return

        if timed_out_message:
            task["status"] = "failed"
            task["error"] = timed_out_message
            _append_log(task, timed_out_message, "error")
            return

        if return_code != 0:
            task["status"] = "failed"
            task["error"] = _topic1_failure_message(payload, output_lines, return_code)
            _append_log(task, task["error"], "error")
            return

        try:
            task["result_rows"] = _parse_project4_result(output_lines) if topic_id == 4 else _resolve_result_rows(task)
            task["status"] = "completed"
            _append_log(task, f"执行完成，共 {len(task['result_rows'])} 条结果", "success")
        except Exception as exc:
            task["status"] = "failed"
            task["error"] = str(exc)
            _append_log(task, f"结果解析失败: {exc}", "error")


def start_task(payload: dict[str, Any]) -> str:
    _cleanup_old_tasks()
    task_id = uuid.uuid4().hex
    task = {
        "id": task_id,
        "status": "running",
        "created_at": time.time(),
        "payload": payload,
        "logs": deque(maxlen=5000),
        "subscribers": [],
        "result_rows": [],
        "error": None,
        "process": None,
    }
    with _lock:
        _tasks[task_id] = task

    target = _run_project3_task if payload.get("topicId") == 3 else _run_process
    thread = threading.Thread(target=target, args=(task_id,), daemon=True)
    thread.start()
    return task_id


def get_task(task_id: str) -> dict[str, Any] | None:
    with _lock:
        task = _tasks.get(task_id)
        if not task:
            return None
        return {
            "taskId": task_id,
            "status": task["status"],
            "error": task["error"],
            "rows": list(task["result_rows"]),
            "logCount": len(task["logs"]),
        }


def subscribe_logs(task_id: str) -> tuple[queue.Queue, dict[str, Any] | None]:
    subscriber: queue.Queue = queue.Queue(maxsize=1000)
    with _lock:
        task = _tasks.get(task_id)
        if not task:
            return subscriber, None
        task["subscribers"].append(subscriber)
        snapshot = list(task["logs"])
        current = {
            "status": task["status"],
            "error": task["error"],
            "rows": list(task["result_rows"]),
        }
    for entry in snapshot:
        subscriber.put(entry)
    return subscriber, current


def unsubscribe_logs(task_id: str, subscriber: queue.Queue) -> None:
    with _lock:
        task = _tasks.get(task_id)
        if not task:
            return
        if subscriber in task["subscribers"]:
            task["subscribers"].remove(subscriber)


def cancel_task(task_id: str) -> bool:
    with _lock:
        task = _tasks.get(task_id)
        if not task:
            return False
        process = task.get("process")
        payload = task.get("payload") or {}
        if process and process.poll() is None:
            process.kill()
        if payload.get("topicId") == 3:
            try:
                _cleanup_project3_server_port(int(payload.get("port") or PROJECT3_DEFAULT_PORT))
            except Exception:
                pass
        if task["status"] == "running":
            task["status"] = "failed"
            task["error"] = "任务已取消"
            _append_log(task, "任务已取消", "error")
        return True
